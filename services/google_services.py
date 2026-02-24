"""
Google Services Module
Handles Google Forms creation and Google Sheets operations
"""
import os
import json
from typing import List, Dict, Any, Optional
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

from config import GOOGLE_CREDENTIALS_FILE, GOOGLE_TOKEN_FILE, GOOGLE_SCOPES


class GoogleServicesManager:
    """Manager class for Google Forms and Sheets operations"""
    
    def __init__(self):
        """Initialize the Google Services Manager"""
        self.creds = None
        self.forms_service = None
        self.sheets_service = None
        self.drive_service = None
        self._authenticate()
    
    def _authenticate(self):
        """Authenticate with Google APIs"""
        # Check for existing token
        if os.path.exists(GOOGLE_TOKEN_FILE):
            self.creds = Credentials.from_authorized_user_file(GOOGLE_TOKEN_FILE, GOOGLE_SCOPES)
        
        # If no valid credentials, get new ones
        if not self.creds or not self.creds.valid:
            if self.creds and self.creds.expired and self.creds.refresh_token:
                self.creds.refresh(Request())
            else:
                if not os.path.exists(GOOGLE_CREDENTIALS_FILE):
                    raise FileNotFoundError(
                        f"Google credentials file not found: {GOOGLE_CREDENTIALS_FILE}\n"
                        "Please download it from Google Cloud Console."
                    )
                flow = InstalledAppFlow.from_client_secrets_file(
                    GOOGLE_CREDENTIALS_FILE, GOOGLE_SCOPES
                )
                self.creds = flow.run_local_server(port=0)
            
            # Save credentials for next run
            with open(GOOGLE_TOKEN_FILE, 'w') as token:
                token.write(self.creds.to_json())
        
        # Build services
        self.forms_service = build('forms', 'v1', credentials=self.creds)
        self.sheets_service = build('sheets', 'v4', credentials=self.creds)
        self.drive_service = build('drive', 'v3', credentials=self.creds)
    
    async def create_google_form(
        self, 
        title: str, 
        questions: List[Dict[str, Any]],
        description: str = ""
    ) -> Dict[str, str]:
        """
        Create a Google Form with the given questions
        
        Args:
            title: Form title
            questions: List of question dictionaries
            description: Optional form description
            
        Returns:
            Dictionary with form ID and URL
        """
        try:
            # Create the form
            form = {
                "info": {
                    "title": title,
                    "documentTitle": title
                }
            }
            
            result = self.forms_service.forms().create(body=form).execute()
            form_id = result['formId']
            
            # Build batch update request for questions
            requests = []
            
            # Add description if provided
            if description:
                requests.append({
                    "updateFormInfo": {
                        "info": {
                            "description": description
                        },
                        "updateMask": "description"
                    }
                })
            
            # Add questions
            for idx, q in enumerate(questions):
                question_item = {
                    "createItem": {
                        "item": {
                            "title": q.get('question_text', f'Question {idx + 1}'),
                            "questionItem": {
                                "question": {
                                    "required": True,
                                    "choiceQuestion": {
                                        "type": "RADIO",
                                        "options": [
                                            {"value": f"{key}: {value}"}
                                            for key, value in q.get('options', {}).items()
                                        ],
                                        "shuffle": False
                                    }
                                }
                            }
                        },
                        "location": {
                            "index": idx
                        }
                    }
                }
                requests.append(question_item)
            
            # Execute batch update
            if requests:
                self.forms_service.forms().batchUpdate(
                    formId=form_id,
                    body={"requests": requests}
                ).execute()
            
            # Get form URL
            form_info = self.forms_service.forms().get(formId=form_id).execute()
            responder_uri = form_info.get('responderUri', f'https://docs.google.com/forms/d/{form_id}/viewform')
            
            return {
                "form_id": form_id,
                "edit_url": f"https://docs.google.com/forms/d/{form_id}/edit",
                "response_url": responder_uri,
                "success": True
            }
            
        except HttpError as e:
            return {
                "error": str(e),
                "success": False
            }
    
    async def create_google_sheet(
        self, 
        title: str, 
        questions: List[Dict[str, Any]],
        predictions: Dict[str, Any]
    ) -> Dict[str, str]:
        """
        Create a Google Sheet with questions and predicted responses
        
        Args:
            title: Spreadsheet title
            questions: List of question dictionaries
            predictions: Predicted responses data
            
        Returns:
            Dictionary with sheet ID and URL
        """
        try:
            # Create spreadsheet
            spreadsheet = {
                "properties": {
                    "title": title
                },
                "sheets": [
                    {"properties": {"title": "Questions"}},
                    {"properties": {"title": "Predicted Responses"}},
                    {"properties": {"title": "Statistics"}}
                ]
            }
            
            result = self.sheets_service.spreadsheets().create(body=spreadsheet).execute()
            spreadsheet_id = result['spreadsheetId']
            
            # Prepare data for Questions sheet
            questions_data = [["#", "Question", "Option A", "Option B", "Option C", "Option D", "Correct Answer", "Difficulty", "Topic"]]
            for i, q in enumerate(questions):
                options = q.get('options', {})
                questions_data.append([
                    i + 1,
                    q.get('question_text', ''),
                    options.get('A', ''),
                    options.get('B', ''),
                    options.get('C', ''),
                    options.get('D', ''),
                    q.get('correct_answer', ''),
                    q.get('difficulty', ''),
                    q.get('topic', '')
                ])
            
            # Prepare data for Predictions sheet
            responses = predictions.get('responses', [])
            predictions_header = ["Respondent ID", "Type", "Score", "Percentage"]
            predictions_header.extend([f"Q{i+1}" for i in range(len(questions))])
            
            predictions_data = [predictions_header]
            for resp in responses:
                row = [
                    resp.get('respondent_id', ''),
                    resp.get('respondent_type', ''),
                    resp.get('score', ''),
                    resp.get('score_percentage', '')
                ]
                answers = resp.get('answers', {})
                for i in range(len(questions)):
                    row.append(answers.get(f'Q{i+1}', ''))
                predictions_data.append(row)
            
            # Prepare Statistics sheet
            stats = predictions.get('statistics', {})
            stats_data = [
                ["Survey Statistics"],
                [""],
                ["Overall Statistics"],
                ["Average Score", stats.get('average_score', 'N/A')],
                ["Average Percentage", f"{stats.get('average_percentage', 'N/A')}%"],
                [""],
                ["Question Performance"],
                ["Question", "Correct Count", "Correct %", "Most Common Wrong Answer", "Wrong Count"]
            ]
            
            q_difficulty = stats.get('question_difficulty', {})
            common_wrong = stats.get('common_wrong_answers', {})
            
            for i in range(len(questions)):
                q_key = f'Q{i+1}'
                q_stats = q_difficulty.get(q_key, {})
                wrong_stats = common_wrong.get(q_key, {})
                stats_data.append([
                    q_key,
                    q_stats.get('correct_count', 0),
                    f"{q_stats.get('percentage', 0)}%",
                    wrong_stats.get('most_common_wrong', 'N/A'),
                    wrong_stats.get('count', 0)
                ])
            
            # Batch update all sheets
            batch_data = [
                {
                    "range": "Questions!A1",
                    "values": questions_data
                },
                {
                    "range": "Predicted Responses!A1",
                    "values": predictions_data
                },
                {
                    "range": "Statistics!A1",
                    "values": stats_data
                }
            ]
            
            self.sheets_service.spreadsheets().values().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={
                    "valueInputOption": "RAW",
                    "data": batch_data
                }
            ).execute()
            
            # Format the sheets (make headers bold)
            format_requests = [
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": 0,
                            "startRowIndex": 0,
                            "endRowIndex": 1
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.2, "green": 0.4, "blue": 0.8},
                                "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}}
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,textFormat)"
                    }
                },
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": 1,
                            "startRowIndex": 0,
                            "endRowIndex": 1
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.2, "green": 0.6, "blue": 0.4},
                                "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}}
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,textFormat)"
                    }
                }
            ]
            
            self.sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": format_requests}
            ).execute()
            
            return {
                "spreadsheet_id": spreadsheet_id,
                "url": f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}",
                "success": True
            }
            
        except HttpError as e:
            return {
                "error": str(e),
                "success": False
            }
    
    def create_local_excel(
        self, 
        filename: str,
        questions: List[Dict[str, Any]],
        predictions: Dict[str, Any]
    ) -> str:
        """
        Create a local Excel file with questions and predictions
        (Fallback when Google Sheets is not available)
        
        Args:
            filename: Output filename
            questions: List of question dictionaries
            predictions: Predicted responses data
            
        Returns:
            Path to created file
        """
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_fill_green = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Questions Sheet
        ws_questions = wb.active
        ws_questions.title = "Questions"
        
        questions_headers = ["#", "Question", "Option A", "Option B", "Option C", "Option D", "Correct Answer", "Difficulty", "Topic"]
        for col, header in enumerate(questions_headers, 1):
            cell = ws_questions.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, q in enumerate(questions, 2):
            options = q.get('options', {})
            row_data = [
                row_idx - 1,
                q.get('question_text', ''),
                options.get('A', ''),
                options.get('B', ''),
                options.get('C', ''),
                options.get('D', ''),
                q.get('correct_answer', ''),
                q.get('difficulty', ''),
                q.get('topic', '')
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws_questions.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
        
        # Adjust column widths
        ws_questions.column_dimensions['B'].width = 50
        for col in ['C', 'D', 'E', 'F']:
            ws_questions.column_dimensions[col].width = 25
        
        # Predictions Sheet
        ws_predictions = wb.create_sheet("Predicted Responses")
        
        predictions_headers = ["Respondent ID", "Type", "Score", "Percentage"]
        predictions_headers.extend([f"Q{i+1}" for i in range(len(questions))])
        
        for col, header in enumerate(predictions_headers, 1):
            cell = ws_predictions.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill_green
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        responses = predictions.get('responses', [])
        for row_idx, resp in enumerate(responses, 2):
            row_data = [
                resp.get('respondent_id', ''),
                resp.get('respondent_type', ''),
                resp.get('score', ''),
                resp.get('score_percentage', '')
            ]
            answers = resp.get('answers', {})
            for i in range(len(questions)):
                row_data.append(answers.get(f'Q{i+1}', ''))
            
            for col, value in enumerate(row_data, 1):
                cell = ws_predictions.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
        
        # Statistics Sheet
        ws_stats = wb.create_sheet("Statistics")
        
        stats = predictions.get('statistics', {})
        
        ws_stats['A1'] = "Survey Statistics"
        ws_stats['A1'].font = Font(bold=True, size=14)
        
        ws_stats['A3'] = "Overall Statistics"
        ws_stats['A3'].font = Font(bold=True)
        ws_stats['A4'] = "Average Score"
        ws_stats['B4'] = stats.get('average_score', 'N/A')
        ws_stats['A5'] = "Average Percentage"
        ws_stats['B5'] = f"{stats.get('average_percentage', 'N/A')}%"
        
        ws_stats['A7'] = "Question Performance"
        ws_stats['A7'].font = Font(bold=True)
        
        stats_headers = ["Question", "Correct Count", "Correct %", "Most Common Wrong", "Wrong Count"]
        for col, header in enumerate(stats_headers, 1):
            cell = ws_stats.cell(row=8, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        q_difficulty = stats.get('question_difficulty', {})
        common_wrong = stats.get('common_wrong_answers', {})
        
        for row_idx, i in enumerate(range(len(questions)), 9):
            q_key = f'Q{i+1}'
            q_stats = q_difficulty.get(q_key, {})
            wrong_stats = common_wrong.get(q_key, {})
            
            row_data = [
                q_key,
                q_stats.get('correct_count', 0),
                f"{q_stats.get('percentage', 0)}%",
                wrong_stats.get('most_common_wrong', 'N/A'),
                wrong_stats.get('count', 0)
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws_stats.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
        
        # Add a chart
        if len(questions) > 0:
            chart = BarChart()
            chart.title = "Correct Answers by Question"
            chart.x_axis.title = "Question"
            chart.y_axis.title = "Correct Count"
            
            data = Reference(ws_stats, min_col=2, min_row=8, max_row=8+len(questions))
            categories = Reference(ws_stats, min_col=1, min_row=9, max_row=8+len(questions))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.shape = 4
            ws_stats.add_chart(chart, "G3")
        
        # Save the workbook
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        wb.save(filename)
        return filename


# Create singleton instance (lazy initialization)
_google_services = None

def get_google_services() -> GoogleServicesManager:
    """Get or create the Google Services Manager instance"""
    global _google_services
    if _google_services is None:
        _google_services = GoogleServicesManager()
    return _google_services
